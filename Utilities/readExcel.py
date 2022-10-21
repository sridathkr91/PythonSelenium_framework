import openpyxl

def getRowCount(path,sheetName):
    wb=openpyxl.load_workbook(path)
    ws=wb[sheetName]
    row=ws.max_row
    return row

def getColumnCount(path,sheetName):
    wb=openpyxl.load_workbook(path)
    ws=wb[sheetName]
    column=ws.max_column
    return column

def read(path,sheetName,row,column):
    wb=openpyxl.load_workbook(path)
    ws=wb[sheetName]
    return ws.cell(row=row,column=column).value

def write(path,sheetName,row,column,data):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheetName]
    ws.cell(row=row,column=column).value=data
    wb.save_workbook(path)
